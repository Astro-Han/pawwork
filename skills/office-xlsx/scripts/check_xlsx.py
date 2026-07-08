#!/usr/bin/env python3
"""Pre-delivery gate for the office-xlsx route: open the finished .xlsx with
openpyxl and verify it really has the structure the task demands (sheets,
real formulas, charts) before the run claims success. A clean OK here means
the workbook opens, the formulas are stored as formulas (not baked-in text),
no formula smells like injected user data, and any required charts exist.

Charts are counted from the package parts (xl/charts/chart*.xml) rather than
openpyxl's private read-back (ws._charts), so the gate cannot false-FAIL on
chart types openpyxl reads back incompletely.

Run through uv so openpyxl resolves from the mirror:
  uv run python <skill>/scripts/check_xlsx.py <book.xlsx> --min-sheets 1 [--require-formula] [--require-chart]
"""

from __future__ import annotations

import argparse
import re
import sys
import zipfile

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        "FAIL: openpyxl is not importable. Run this through 'uv run python ...' from the "
        "directory that holds the office-xlsx pyproject.toml (uv must be on PATH).",
        file=sys.stderr,
    )
    sys.exit(2)

# Formula-injection tripwires. Generated models never need these: external
# fetch / DDE / process launch have no place in a deliverable, and a real
# hyperlink should use cell.hyperlink (an attribute), never the HYPERLINK()
# formula — a HYPERLINK formula usually means user data leaked in as a formula.
DANGEROUS_FORMULA = re.compile(
    r"\b(WEBSERVICE|IMPORTXML|IMPORTDATA|IMPORTHTML|IMPORTFEED|IMPORTRANGE|HYPERLINK|EXEC|DDE)\s*\(|\|",
    re.IGNORECASE,
)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx", help="path to the built .xlsx")
    parser.add_argument("--min-sheets", type=int, default=1, help="minimum worksheet count")
    parser.add_argument("--require-formula", action="store_true", help="need at least one real formula cell")
    parser.add_argument("--require-chart", action="store_true", help="need at least one chart part in the package")
    args = parser.parse_args()

    try:
        wb = load_workbook(args.xlsx)
    except Exception as error:  # noqa: BLE001 - report any open failure as a gate failure
        print(f"FAIL: cannot open {args.xlsx}: {error}")
        return 1

    # Count chart parts at the package level (stable OOXML structure, no
    # reliance on openpyxl's private ws._charts read-back).
    with zipfile.ZipFile(args.xlsx) as zf:
        chart_count = sum(1 for n in zf.namelist() if re.fullmatch(r"xl/charts/chart\d+\.xml", n))

    sheets = wb.sheetnames
    formula_cells = 0
    populated_cells = 0
    dangerous: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                populated_cells += 1
                # When loading from file, openpyxl marks every formula cell with
                # data_type "f". A string cell (data_type "s") that merely starts
                # with "=" was force-stored as text — that is the SAFE state for
                # user data, so it must count as text, not formula.
                if cell.data_type == "f":
                    formula_cells += 1
                    if isinstance(cell.value, str) and DANGEROUS_FORMULA.search(cell.value):
                        dangerous.append(f"{ws.title}!{cell.coordinate}: {cell.value[:60]}")

    problems: list[str] = []
    if len(sheets) < args.min_sheets:
        problems.append(f"expected at least {args.min_sheets} sheet(s), found {len(sheets)}")
    if populated_cells == 0:
        problems.append("workbook has no populated cells")
    if args.require_formula and formula_cells < 1:
        problems.append("task needs real formulas but no formula cell was found (values may be hard-coded)")
    if args.require_chart and chart_count < 1:
        problems.append("task needs a chart but no xl/charts/chart*.xml part was found")
    for hit in dangerous:
        problems.append(
            f"suspicious formula (injection tripwire) at {hit} — user-supplied data starting with "
            "=/+/-/@ must be stored as text; links go through cell.hyperlink, not HYPERLINK()"
        )

    for problem in problems:
        print(f"FAIL: {problem}")
    if problems:
        print(f"FAIL: {len(problems)} structure gap(s). Fix the generator and rebuild; do not hand-edit the zip.")
        return 1
    print(
        f"OK: {len(sheets)} sheet(s), {populated_cells} populated cell(s), "
        f"{formula_cells} formula cell(s), {chart_count} chart part(s) — workbook matches the brief."
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
